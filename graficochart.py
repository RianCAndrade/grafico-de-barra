from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#Importando a planilha
plan = load_workbook('tabelaexcel') # tabela excel 
planl = plan['nomedaplanilha']

#Referencias de linhas e colunas
min_column = plan.active.min_column
max_column = plan.active.max_column 
min_row = plan.active.min_row
max_row = plan.active.max_row

barchart = BarChart()

#Pegando os dados e as categorias
dados = Reference(
    planl,
    min_col = min_column + 1,
    max_col = max_column,
    min_row = min_row,
    max_row = max_row
)

categorias = Reference(
    planl,
    min_col = min_column,
    max_col = min_column,
    min_row = min_row + 1,
    max_row = max_row
)

barchart.add_data(dados, titles_from_data=True)
barchart.set_categories(categorias)

planl.add_chart(barchart, 'B10')
barchart.title = 'titulo'
barchart.style = 2

plan.save('nomedanovaplanilha') #salvando o grafico da planilha excel